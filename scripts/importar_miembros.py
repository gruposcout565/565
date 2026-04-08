#!/usr/bin/env python3
"""
Importa protagonistas desde el Excel exportado de la organización scout.
Uso: python3 scripts/importar_miembros.py [--dry-run]

Columnas usadas del Excel:
  Documento      → dni
  Nombre         → apellido + nombre  (formato "Apellido, Nombre")
  Fecha Nacimiento → fecha_nacimiento
  Calle          → direccion + barrio (extrae "B° ..." como barrio)
  Localidad      → barrio (si no se encontró barrio en la calle)
  Telefono       → telefono
  Email          → email
  Religion       → religion
  Rama           → rama
  Fecha Primer Afiliacion → fecha_ingreso
"""

import sys
import json
import re
import urllib.request
import urllib.error
from datetime import datetime

import openpyxl

# ── Configuración ────────────────────────────────────────────────────────────

EXCEL_PATH = '/Users/franciscomansilla/Downloads/miembros (1).xlsx'

SUPABASE_URL = 'https://aebnchyntggsyfzfbomt.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImFlYm5jaHludGdnc3lmemZib210Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MzcwMzk1NCwiZXhwIjoyMDg5Mjc5OTU0fQ.k17y_RmQBUJm08GbClfcM99T87r-QH3ZqN4LfqCt_Io'

DRY_RUN = '--dry-run' in sys.argv

# ── Helpers ──────────────────────────────────────────────────────────────────

def parse_date(value: str | None) -> str | None:
    """Convierte 'DD/MM/YYYY' a 'YYYY-MM-DD'."""
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), '%d/%m/%Y').strftime('%Y-%m-%d')
    except ValueError:
        return None


def split_nombre(raw: str) -> tuple[str, str]:
    """
    Separa "Apellido Apellido, Nombre Nombre" en (nombre, apellido).
    Si no hay coma, pone todo en apellido.
    """
    if ',' in raw:
        apellido, nombre = raw.split(',', 1)
        return nombre.strip().title(), apellido.strip().title()
    return '', raw.strip().title()


def parse_direccion(calle: str | None) -> tuple[str | None, str | None]:
    """
    Extrae dirección y barrio de un campo como:
      "General Pedernera 692 B° San Vicente"
    Devuelve (direccion_sin_barrio, barrio) o (calle_completa, None).
    """
    if not calle:
        return None, None
    match = re.search(r'\bB[°º]\s*(.+)$', calle, re.IGNORECASE)
    if match:
        barrio = match.group(1).strip().title()
        direccion = calle[:match.start()].strip().rstrip(',').strip()
        return direccion or None, barrio
    return calle.strip(), None


def supabase_request(method: str, path: str, body=None) -> dict:
    url = f'{SUPABASE_URL}/rest/v1/{path}'
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(
        url,
        data=data,
        method=method,
        headers={
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal',
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read()
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        body_err = e.read().decode()
        raise RuntimeError(f'HTTP {e.code}: {body_err}')


def get_existing_dnis() -> set[str]:
    url = f'{SUPABASE_URL}/rest/v1/beneficiarios?select=dni'
    req = urllib.request.Request(
        url,
        headers={
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
        },
    )
    with urllib.request.urlopen(req) as resp:
        rows = json.loads(resp.read())
    return {str(r['dni']) for r in rows if r.get('dni')}


# ── Lectura del Excel ────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active
headers = [cell.value for cell in ws[1]]

def col(row, name):
    idx = headers.index(name)
    v = row[idx]
    return str(v).strip() if v is not None else None

# ── Proceso principal ────────────────────────────────────────────────────────

print(f'{"[DRY RUN] " if DRY_RUN else ""}Leyendo {ws.max_row - 1} filas del Excel...')
print('Obteniendo DNIs existentes en Supabase...')
existing_dnis = get_existing_dnis() if not DRY_RUN else set()
print(f'  → {len(existing_dnis)} protagonistas ya registrados.\n')

inserted = 0
skipped = 0
errors = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    nombre_raw = col(row, 'Nombre')
    if not nombre_raw:
        continue

    nombre, apellido = split_nombre(nombre_raw)
    dni = col(row, 'Documento')
    rama = col(row, 'Rama')
    calle = col(row, 'Calle')
    localidad = col(row, 'Localidad')

    direccion, barrio = parse_direccion(calle)
    # Si no se encontró barrio en la calle, usar localidad
    if not barrio and localidad and localidad.lower() not in ('cordoba', 'córdoba'):
        barrio = localidad.strip().title()

    record = {
        'nombre': nombre,
        'apellido': apellido,
        'dni': dni,
        'fecha_nacimiento': parse_date(col(row, 'Fecha Nacimiento')),
        'telefono': col(row, 'Telefono'),
        'email': col(row, 'Email'),
        'religion': col(row, 'Religion'),
        'direccion': direccion,
        'barrio': barrio,
        'rama': rama,
        'fecha_ingreso': parse_date(col(row, 'Fecha Primer Afiliacion')) or datetime.today().strftime('%Y-%m-%d'),
        'tipo_cuota': 'mensual',
        'activo': True,
    }

    label = f'{apellido}, {nombre} (DNI: {dni})'

    if dni and dni in existing_dnis:
        print(f'  SKIP   {label}  — ya existe')
        skipped += 1
        continue

    if DRY_RUN:
        print(f'  DRY    {label}')
        print(f'         rama={rama} | dir={direccion} | barrio={barrio} | nacimiento={record["fecha_nacimiento"]}')
        inserted += 1
        continue

    try:
        supabase_request('POST', 'beneficiarios', record)
        print(f'  OK     {label}')
        inserted += 1
    except RuntimeError as e:
        print(f'  ERROR  {label}: {e}')
        errors += 1

print(f'\n{"─"*50}')
print(f'Insertados : {inserted}')
print(f'Saltados   : {skipped}  (ya existían)')
print(f'Errores    : {errors}')
